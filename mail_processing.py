import asyncio
import aiofiles
import pandas as pd
import aiohttp
from base64 import urlsafe_b64encode
from datetime import datetime, date, timedelta
from googleapiclient.errors import HttpError
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from base64 import urlsafe_b64decode
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from mimetypes import guess_type as guess_mime_type
from openpyxl import load_workbook
import os
import xlrd
import pytz

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly", 
    "https://www.googleapis.com/auth/gmail.send",
]

class GmailService:
    def __init__(self, token_file="token.json", credentials_file="credentials.json"):
        self.service = self.authenticate(token_file, credentials_file)

    def authenticate(self, token_file, credentials_file):
        creds = None
        if os.path.exists(token_file):
            creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(credentials_file, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_file, "w") as token:
                token.write(creds.to_json())
        return build('gmail', 'v1', credentials=creds)

    async def search_messages(self, query):
        result = self.service.users().messages().list(userId='me', q=query).execute()
        messages = result.get('messages', [])
        while 'nextPageToken' in result:
            page_token = result['nextPageToken']
            result = self.service.users().messages().list(userId='me', q=query, pageToken=page_token).execute()
            messages.extend(result.get('messages', []))
        return messages
    
    # Send email
    def _build_simple_message(self, to: str, subject: str, body_text: str = "", body_html: str = None) -> dict:
        """Builds a simple Gmail message (text or HTML, no attachments)."""
        if body_html:
            msg = MIMEMultipart("alternative")
            msg.attach(MIMEText(body_text or "", "plain", "utf-8"))
            msg.attach(MIMEText(body_html, "html", "utf-8"))
        else:
            msg = MIMEText(body_text or "", "plain", "utf-8")

        msg["To"] = to
        msg["Subject"] = subject
        raw_message = urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
        return {"raw": raw_message}

    def send_email(self, to: str, subject: str, body_text: str = "", body_html: str = None):
        """Send an email (synchronously)."""
        message = self._build_simple_message(to, subject, body_text, body_html)
        try:
            return (
                self.service.users()
                .messages()
                .send(userId="me", body=message)
                .execute()
            )
        except HttpError as e:
            raise RuntimeError(f"Gmail send failed: {e}") from e
        
    async def send_email_async(self, to: str, subject: str, body_text: str = "", body_html: str = None):
        """Async wrapper to send emails without blocking the event loop."""
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, lambda: self.send_email(to, subject, body_text, body_html))

    async def email_georgi(self, subject: str, body_text: str = "", body_html: str = None):
        """Convenience method to email georgi.manev@entra.energy."""
        return await self.send_email_async(
            to="georgi.manev@entra.energy",
            subject=subject,
            body_text=body_text,
            body_html=body_html,
        )
    
    async def email_rali(self, subject: str, body_text: str = "", body_html: str = None):
        """Convenience method to email ralitsa.rumenova@entra.energy."""
        return await self.send_email_async(
            to="ralitsa.rumenova@entra.energy",
            subject=subject,
            body_text=body_text,
            body_html=body_html,
        )

    async def parse_parts(self, service, parts, folder_name, message):
        if parts:
            for part in parts:
                filename = part.get("filename")
                mimeType = part.get("mimeType")
                body = part.get("body")
                data = body.get("data")
                file_size = body.get("size")
                part_headers = part.get("headers")

                if part.get("parts"):
                    await self.parse_parts(service, part.get("parts"), folder_name, message)
                else:
                    for part_header in part_headers:
                        part_header_name = part_header.get("name")
                        part_header_value = part_header.get("value")
                        if part_header_name == "Content-Disposition":
                            if "attachment" in part_header_value:
                                attachment_id = body.get("attachmentId")
                                attachment = service.users().messages().attachments().get(id=attachment_id, userId='me', messageId=message['id']).execute()
                                data = attachment.get("data")
                                filepath = os.path.join(folder_name, filename)
                                if data:
                                    async with aiofiles.open(filepath, "wb") as f:
                                        await f.write(urlsafe_b64decode(data))

    async def read_message(self, message, price_clearing=False):
        msg = self.service.users().messages().get(userId='me', id=message['id'], format='full').execute()
        payload = msg['payload']
        headers = payload.get("headers")
        parts = payload.get("parts")
        folder_name = "email"
        mail_hour = None
        if headers:
            for header in headers:
                if header.get("name").lower() == "subject":
                    folder_name = "enProMail"
                elif header.get("name").lower() == "date":
                    date = header.get("value")
                    local_tz = pytz.timezone('Europe/Sofia')
                    date_obj = datetime.strptime(date, "%a, %d %b %Y %H:%M:%S %z")
                    date_obj = date_obj.astimezone(local_tz)
                    mail_hour = date_obj.hour
               
        print(f"mail_hour:{mail_hour}")
        if price_clearing:
            if mail_hour and mail_hour >=13: #Filter additional mails with clearings from EnPro
                await self.parse_parts(self.service, parts, folder_name, message)
                print("=" * 50)
        else:
            await self.parse_parts(self.service, parts, folder_name, message)
            print("=" * 50)

class FileManager:

    def __init__(self, farm) -> None:
        self.farm = farm


    def get_file_name(self, folder):
        today = date.today()
        d1 = today.strftime("%d.%m.%Y")
        st = folder.split("_")[1].split("xls")[0]
        file_date = st.split('.')
        d = file_date[0]
        m = file_date[1]
        y = file_date[2]
        name_date = d + "." + m + "." + y
        #print(f"Name Date: {name_date} || {d1}")
        return name_date == d1

    async def process_files(self):
        """
        Scans enProMail for .xls or .xlsx files, reads the first matching file,
        and returns the forecast based on self.farm ('neykovo' or 'aris').
        Requires: xlrd (for .xls), openpyxl (for .xlsx)
        """
        def _open_first_sheet(filepath):
            ext = os.path.splitext(filepath)[1].lower()
            if ext == ".xls":
                wb = xlrd.open_workbook(filepath)
                return ("xls", wb.sheet_by_index(0))
            elif ext == ".xlsx":
                wb = load_workbook(filepath, data_only=True, read_only=True)
                return ("xlsx", wb.worksheets[0])
            else:
                raise ValueError(f"Unsupported extension: {ext}")

        def _cell_value(sheet_kind, sheet, r, c):
            # r,c are 0-based like your original code
            if sheet_kind == "xls":
                return sheet.cell_value(r, c)
            else:  # xlsx via openpyxl
                return sheet.cell(row=r + 1, column=c + 1).value

        def _excel_date_to_date(val):
            # Handles both engines: datetime or Excel serial number
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, (int, float)):
                # Use xlrd's converter with datemode=0 (Excel 1900 system)
                return xlrd.xldate_as_datetime(val, 0).date()
            # Fallback: try pandas to_datetime
            try:
                return pd.to_datetime(val).date()
            except Exception:
                raise ValueError(f"Unrecognized date cell value: {val!r}")

        fn = "enProMail"
        for root, dirs, files in os.walk(fn):
            # accept both .xls and .xlsx
            excel_files = [f for f in files if f.lower().endswith((".xls", ".xlsx"))]
            for exfile in excel_files:
                
                my_file = self.get_file_name(exfile)
                if not my_file:
                    continue

                filepath = os.path.join(root, exfile)
                kind, sheet = _open_first_sheet(filepath)

                # Read header cells (same coordinates as your original code)
                xl_asset = _cell_value(kind, sheet, 4, 0)  # not used below, but preserved
                raw_date = _cell_value(kind, sheet, 1, 1)
                xl_date = _excel_date_to_date(raw_date)
                xl_date_time = f"{xl_date}T01:15:00"

                # Collect quarter-hour powers from row 4 and 5, starting at column 3 + i
                testlist = []
                neykovo_list = []
                period = (24 * 4) + 1  # 97
                i = 1
                while i < period:
                    i += 1  # i goes 2..96 (kept identical to your logic)
                    col = 3 + i
                    xl_power = _cell_value(kind, sheet, 4, col)          # row 5 (0-based index 4)
                    xl_neykovo_power = _cell_value(kind, sheet, 5, col)  # row 6 (0-based index 5)
                    testlist.append(xl_power)
                    neykovo_list.append(xl_neykovo_power)

                timeIndex = pd.date_range(start=xl_date_time, periods=period - 1, freq="15T")
                dfA = pd.DataFrame(testlist, index=timeIndex, columns=["pow"])                
                dfNeykovo = pd.DataFrame(neykovo_list, index=timeIndex, columns=["pow"])                
                dfA.index.name = "Aris foreacast"
                dfNeykovo.index.name = "Power forecast"               

                if self.farm == "neykovo":                    
                    forecast = await self.forecast_extractor(dfNeykovo)
                elif self.farm == "aris":
                    forecast = await self.forecast_extractor(dfA)
                else:
                    forecast = None                
                
                return forecast  # preserve your early-return behavior

        # If no matching file found:
        return None
                    
    async def forecast_extractor(self, wind_farm_df):                    
        for row in wind_farm_df.itertuples():
            
            timenow = datetime.now()
            quarter_min = self.lookup_quarterly(timenow.minute)                                    
            if quarter_min == 0:
                quarter_hour = timenow.hour + 1
            else:
                quarter_hour = timenow.hour                     
            forecast_hour = row.Index.hour
            forecast_min = row.Index.minute
            if quarter_hour == forecast_hour and forecast_min == quarter_min:                            
                power = row.pow
                print(f"forecast_hour={forecast_hour}:{forecast_min} || quarter_hour={quarter_hour}:{quarter_min} || Real Time:{timenow.hour}:{timenow.minute} || Power:{power}")                
                return power
                    
                   
    
    def lookup_quarterly(self, minutes):
        
        if 0 <= minutes <= 14:
            return 15
        elif 15 <= minutes <= 29:
            return 30
        elif 30 <= minutes <= 44:
            return 45
        elif 45 <= minutes <= 59:
            return 0
        else:
            raise ValueError("Minutes must be between 0 and 59")
   

class ForecastProcessor:
    def __init__(self):        
        self.gmail_service = GmailService()        

    # Call it with clearing to check for clearing mails
    async def proceed_forecast(self, clearing=False):
        now = datetime.now() 
        #temp = datetime.now() - timedelta(days=1)
        after_date = now.strftime("%Y/%m/%d")
        #before_date = temp.strftime("%Y/%m/%d")
        sender_email = "trading@energo-pro.bg"
        query_str = f"from:{sender_email} after:{after_date} "
        print(query_str)
        results = await self.gmail_service.search_messages(query_str)
        print(f"Found {len(results)} results.")        
        for msg in results:
            await self.gmail_service.read_message(msg, price_clearing=clearing)

   
if __name__ == "__main__":
   
    processor = ForecastProcessor()
    file_manager = FileManager("neykovo")
    loop = asyncio.get_event_loop()
    loop.run_until_complete(processor.proceed_forecast(clearing=False))
    loop.run_until_complete(file_manager.process_files())
    
